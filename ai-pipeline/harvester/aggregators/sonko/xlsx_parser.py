"""Parser for SONKO registry XLSX from data.economy.gov.ru.

Downloads (or reads local) the XLSX and parses into SONKOEntry list.
The file is ~9 MB with ~111K rows (title on row 1, headers on row 2, data from row 3+).
"""

import os
from pathlib import Path
from typing import Optional

import structlog

from aggregators.sonko.models import SONKOEntry

logger = structlog.get_logger(__name__)

SONKO_XLSX_URL = (
    "https://data.economy.gov.ru/files/sonko_organizations.xlsx"
)

EXPECTED_HEADERS = [
    "ИНН организации",
    "Наименование организации",
    "Наименование организации",
    "Адрес регистрации",
    "ОГРН",
]


def download_xlsx(dest_path: str, url: str = SONKO_XLSX_URL) -> str:
    """Download SONKO registry XLSX to dest_path."""
    import httpx

    dest = Path(dest_path)
    dest.parent.mkdir(parents=True, exist_ok=True)

    logger.info("Downloading SONKO XLSX", url=url, dest=str(dest))
    with httpx.Client(
        timeout=120, follow_redirects=True, verify=False,
        headers={"User-Agent": "Mozilla/5.0 (Navigator Harvester)"},
    ) as client:
        resp = client.get(url)
        resp.raise_for_status()
        dest.write_bytes(resp.content)

    size_mb = dest.stat().st_size / (1024 * 1024)
    logger.info("SONKO XLSX downloaded", size_mb=f"{size_mb:.1f}")
    return str(dest)


def parse_xlsx(
    path: str,
    limit: Optional[int] = None,
) -> list[SONKOEntry]:
    """Parse SONKO registry XLSX into SONKOEntry list.

    The XLSX has a title in row 1, headers in row 2, and some empty rows
    before data begins (~row 21). Rows without a valid INN are skipped.
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError(
            "openpyxl is required for XLSX parsing. "
            "Install with: pip install openpyxl"
        )

    if not os.path.exists(path):
        raise FileNotFoundError(f"SONKO XLSX not found: {path}")

    logger.info("Parsing SONKO XLSX", path=path, limit=limit)

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    headers: list[str] = []
    for row in ws.iter_rows(min_row=2, max_row=2, values_only=True):
        headers = [_normalize_header(h) for h in row]

    _validate_headers(headers)

    entries: list[SONKOEntry] = []
    errors = 0
    parsed = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True)):
        inn = _clean_str(row[0])
        if not inn or len(inn) < 5:
            continue

        if limit and parsed >= limit:
            break

        try:
            entry = _row_to_entry(row)
            entries.append(entry)
            parsed += 1
        except Exception as e:
            errors += 1
            if errors <= 5:
                logger.warning(
                    "Failed to parse SONKO row",
                    row=row_idx + 3,
                    error=str(e),
                )

    wb.close()

    logger.info(
        "SONKO XLSX parsed",
        total_entries=len(entries),
        errors=errors,
    )
    return entries


def _normalize_header(h: object) -> str:
    if h is None:
        return ""
    return " ".join(str(h).split())


def _validate_headers(headers: list[str]) -> None:
    for i, expected in enumerate(EXPECTED_HEADERS):
        actual = headers[i] if i < len(headers) else "<missing>"
        if expected.lower() not in actual.lower():
            raise ValueError(
                f"Column {i} mismatch: expected '{expected}', got '{actual}'. "
                "SONKO data format may have changed."
            )


def _clean_str(v: object) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _row_to_entry(row: tuple) -> SONKOEntry:
    return SONKOEntry(
        inn=row[0],
        full_name=row[1],
        short_name=row[2] if len(row) > 2 else None,
        address=row[3] if len(row) > 3 else "",
        ogrn=row[4] if len(row) > 4 else "",
        legal_form=row[5] if len(row) > 5 else None,
        okved=row[6] if len(row) > 6 else "",
        sonko_status=row[7] if len(row) > 7 else None,
        nco_status=row[8] if len(row) > 8 else None,
        inclusion_criterion=row[9] if len(row) > 9 else None,
        authority_name=row[10] if len(row) > 10 else None,
        decision_date=row[11] if len(row) > 11 else None,
        inclusion_date=row[12] if len(row) > 12 else None,
    )
