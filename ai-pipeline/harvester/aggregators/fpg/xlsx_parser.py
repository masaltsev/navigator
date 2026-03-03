"""Parser for FPG open data XLSX file.

Downloads (or reads local) the XLSX from президентскиегранты.рф/public/open-data
and parses it into a list of FPGProject models.

The file is ~28 MB with ~188K rows (all contests since 2017).
Uses openpyxl in read_only mode for memory efficiency.
"""

import os
from pathlib import Path
from typing import Optional

import structlog

from aggregators.fpg.models import FPGProject

logger = structlog.get_logger(__name__)

FPG_XLSX_URL = (
    "https://xn--80asg7a0b.xn--80afcdbalict6afooklqi5o.xn--p1ai"
    "/81cf9ab5-07f9-4272-a058-ec578b4e4c61"
)

EXPECTED_HEADERS = [
    "Номер заявки",
    "Конкурс",
    "Организация",
    "ОГРН",
    "ИНН",
    "Регион",
    "Название проекта",
    "Грантовое направление",
]


def download_xlsx(dest_path: str, url: str = FPG_XLSX_URL) -> str:
    """Download FPG open data XLSX to dest_path. Returns the path."""
    import httpx

    dest = Path(dest_path)
    dest.parent.mkdir(parents=True, exist_ok=True)

    logger.info("Downloading FPG XLSX", url=url, dest=str(dest))
    with httpx.Client(
        timeout=300, follow_redirects=True, verify=False,
        headers={"User-Agent": "Mozilla/5.0 (Navigator Harvester)"},
    ) as client:
        resp = client.get(url)
        resp.raise_for_status()
        dest.write_bytes(resp.content)

    size_mb = dest.stat().st_size / (1024 * 1024)
    logger.info("FPG XLSX downloaded", size_mb=f"{size_mb:.1f}")
    return str(dest)


def parse_xlsx(
    path: str,
    limit: Optional[int] = None,
) -> list[FPGProject]:
    """Parse FPG open data XLSX into FPGProject list.

    Args:
        path: Path to the XLSX file.
        limit: Max number of rows to parse (for testing).

    Returns:
        List of FPGProject models.
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError(
            "openpyxl is required for XLSX parsing. "
            "Install with: pip install openpyxl"
        )

    if not os.path.exists(path):
        raise FileNotFoundError(f"FPG XLSX not found: {path}")

    logger.info("Parsing FPG XLSX", path=path, limit=limit)

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    headers: list[str] = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = [_normalize_header(h) for h in row]

    _validate_headers(headers)

    projects: list[FPGProject] = []
    errors = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if limit and row_idx >= limit:
            break

        try:
            project = _row_to_project(row)
            projects.append(project)
        except Exception as e:
            errors += 1
            if errors <= 5:
                logger.warning(
                    "Failed to parse row",
                    row=row_idx + 2,
                    error=str(e),
                )

    wb.close()

    logger.info(
        "FPG XLSX parsed",
        total_rows=len(projects),
        errors=errors,
    )
    return projects


def _normalize_header(h: object) -> str:
    """Normalize header: strip whitespace and newlines."""
    if h is None:
        return ""
    return " ".join(str(h).split())


def _validate_headers(headers: list[str]) -> None:
    """Verify the XLSX has expected columns."""
    for i, expected in enumerate(EXPECTED_HEADERS):
        actual = headers[i] if i < len(headers) else "<missing>"
        if expected.lower() not in actual.lower():
            raise ValueError(
                f"Column {i} mismatch: expected '{expected}', got '{actual}'. "
                "FPG data format may have changed."
            )


def _row_to_project(row: tuple) -> FPGProject:
    """Convert a raw XLSX row tuple to FPGProject."""
    return FPGProject(
        application_number=str(row[0] or "").strip(),
        contest=str(row[1] or "").strip(),
        organization_name=str(row[2] or "").strip(),
        ogrn=row[3],
        inn=row[4],
        region=str(row[5] or "").strip(),
        project_title=str(row[6] or "").strip(),
        grant_direction=str(row[7] or "").strip(),
        budget_requested=_to_float(row[8]),
        budget_total=_to_float(row[9]),
        start_date=row[10],
        end_date=row[11],
        status=str(row[12] or "").strip(),
        grant_decision_date=row[13] if len(row) > 13 else None,
        grant_amount=_to_float(row[14]) if len(row) > 14 else None,
        evaluation=str(row[15]).strip() if len(row) > 15 and row[15] else None,
        violations=str(row[16]).strip() if len(row) > 16 and row[16] else None,
    )


def _to_float(v: object) -> Optional[float]:
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None
